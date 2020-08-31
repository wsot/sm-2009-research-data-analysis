import itertools
import typing as t
from dataclasses import dataclass

import openpyxl

from lib import structures


@dataclass(frozen=True)
class Offset:
    cols: int
    rows: int


@dataclass(frozen=True)
class Cell:
    col: int
    row: int

    def __str__(self) -> str:
        return f"{openpyxl.utils.cell.get_column_letter(self.col)}{self.row}"

    def __add__(self, offset: t.Union[Offset, t.Tuple[int, int]]) -> "Cell":
        if isinstance(offset, tuple) and len(offset) == 2:
            offset = Offset(offset[0], offset[1])
        elif not isinstance(offset, Offset):
            raise TypeError("Only an Offset or an [int, int] tuple can be added to a Cell")
        return Cell(self.col + offset.cols, self.row + offset.rows)

    def __sub__(self, offset: t.Union[Offset, t.Tuple[int, int]]) -> "Cell":
        if isinstance(offset, tuple) and len(offset) == 2:
            offset = Offset(offset[0], offset[1])
        elif not isinstance(offset, Offset):
            raise TypeError("Only an Offset or an [int, int] tuple can be subtracted from a Cell")
        if offset.cols >= self.col or offset.rows >= self.row:
            raise ValueError("Subtracting offset would results in column index < A or row index < 1")
        return Cell(self.col - offset.cols, self.row - offset.rows)


def write_row(values: t.Iterable[str], worksheet: openpyxl.worksheet.worksheet.Worksheet, origin: Cell) -> Cell:
    for value in values:
        worksheet[str(origin)] = value
        origin = origin + Offset(1, 0)
    return origin


def write_cells(
    values: t.Iterable[t.Iterable[str]], worksheet: openpyxl.worksheet.worksheet.Worksheet, origin: Cell
) -> Cell:
    for column_values in values:
        write_row(column_values, worksheet, origin)
        origin = origin + Offset(0, 1)
    return origin


def write_session_header(
    session: structures.Session, worksheet: openpyxl.worksheet.worksheet.Worksheet, origin: Cell
) -> Cell:
    #     worksheet[str(origin)] = "ABC"
    return origin + Offset(0, 1)


def write_trial_header(
    trial: structures.Trial, worksheet: openpyxl.worksheet.worksheet.Worksheet, origin: Cell
) -> Cell:
    assert isinstance(trial, structures.AcousticTrial)
    return write_cells(
        (
            ("Trial", str(trial.trial_number)),
            ("Base frequency", str(trial.base_frequency)),
            ("Alternate frequency", str(trial.alternate_frequency)),
            ("Trial start timestamp", str(trial.start_timestamp)),
            ["Amplitudes"] + [str(a) for a in trial.amplitudes],
        ),
        worksheet,
        origin,
    )


def write_trial_data(trial: structures.Trial, worksheet: openpyxl.worksheet.worksheet.Worksheet, origin: Cell) -> Cell:
    assert isinstance(trial, structures.IncludedAcousticTrial)
    write_row(
        itertools.chain(iter(("Frequency",)), (str(t.cast(structures.Tone, s).frequency) for s in trial.stimuli)),
        worksheet,
        origin,
    )
    origin += Offset(0, 1)
    write_row(
        itertools.chain(iter(("Attenuation",)), (str(t.cast(structures.Tone, s).attenuation) for s in trial.stimuli)),
        worksheet,
        origin,
    )
    origin += Offset(0, 1)
    write_row(
        (str(round(t.stimulus_start_relative_timestamp, 2)) for t in trial.stimuli), worksheet, origin + Offset(1, 0)
    )
    origin += Offset(0, 1)
    write_row(("Channel", "During stimulus"), worksheet, origin)
    origin += Offset(0, 1)
    # Create channel numbers
    write_cells(
        [[str(channel_idx + 1)] for channel_idx in range(trial.in_stimulus_spike_counts.shape[1])], worksheet, origin
    )
    origin = write_cells(trial.in_stimulus_spike_counts.transpose(), worksheet, origin + Offset(1, 0))
    origin -= Offset(1, 0)
    write_row(("Channel", "During inter-stimulus interval"), worksheet, origin)
    origin += Offset(0, 1)
    # Create channel numbers
    write_cells(
        [[str(channel_idx + 1)] for channel_idx in range(trial.in_stimulus_spike_counts.shape[1])], worksheet, origin
    )
    origin = write_cells(trial.out_stimulus_spike_counts.transpose(), worksheet, origin + Offset(1, 0))
    origin += Offset(-1, 1)
    return origin


def write_everything(session_results: structures.Session, destination_filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    offset = Cell(1, 1)
    ws.column_dimensions[openpyxl.utils.cell.get_column_letter(1)].width = 20
    write_session_header(session_results, ws, offset)
    for trial in session_results.trials:
        offset = write_trial_header(trial, ws, offset)
        offset = write_trial_data(trial, ws, offset)

    wb.save(destination_filename)
